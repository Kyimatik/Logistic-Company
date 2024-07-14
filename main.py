import json
import asyncio
import logging
import sys

import time 
from aiogram import Bot, Dispatcher, Router, types , F 
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart , Command
from aiogram.types import Message , CallbackQuery ,FSInputFile
from aiogram.utils.markdown import hbold
from aiogram.fsm.context import FSMContext
from aiogram.types import ReplyKeyboardMarkup , InlineKeyboardMarkup , InlineKeyboardButton ,ReplyKeyboardRemove 
from states import Newlog
from states import Change
from states import Calc
from states import sendall
import buttons

import os
import time
from openpyxl import Workbook, load_workbook


admin_id = 6355200375
TOKEN = "7352009164:AAHD84f31ubJwRvt9JX8AIuXeaSaY7bOSBw"

bot = Bot(TOKEN)
dp = Dispatcher()
router = Router()
#Загрузка данных из базы данных usersdb
def load_data():
    try:
        with open('usersdb.json', 'r') as json_file:
            data = json.load(json_file)
    except FileNotFoundError:
        data = {}
    return data

def save_data(data):
    with open('usersdb.json', 'w') as json_file:
        json.dump(data, json_file, indent=4)









@dp.message(Command("start"))
async def start(message: Message , state: FSMContext):
    data = load_data()
    if str(message.from_user.id) in data:
        await message.answer("Вы уже зарегистрированы!",reply_markup=buttons.mainkb)
    else:
       await message.answer(f"""Привет <b>{message.from_user.first_name}</b>!
Мы компания <i>NicolaLogistics</i>
Для дальнейшей работы нам нужно вас зарегистрировать.
Для начала как вас зовут ? 
    """,parse_mode="HTML")
       await state.set_state(Newlog.Name)
        


@dp.message(Newlog.Name)
async def getnewusername(message: Message, state: FSMContext):
    await state.update_data(Name=message.text)
    await state.set_state(Newlog.Surname)
    await message.answer("Отлично! Теперь напишите вашу фамилию.")

@dp.message(Newlog.Surname)
async def getnewsurname(message: Message, state: FSMContext):
    await state.update_data(Surname=message.text)
    await state.set_state(Newlog.Number)
    await message.answer("С фамиилей разобрались теперь ваш рабочий номер телефона!")

@dp.message(Newlog.Number)
async def getnewnubmber(message: Message , state: FSMContext):
    await state.update_data(Number=message.text)
    user_id = message.from_user.id
    data = await state.get_data()
    await state.clear()
    await message.answer("""Отлично, в таком случае регистрация завершена 🔥🔥🔥""",reply_markup=buttons.mainkb)
    data1 = load_data()

    # Добавляем информацию о новом пользователе в словарь данных
    data1[str(user_id)] = {
        "name": str(data["Name"]),
        "surname": data["Surname"],
        "phone_number": data["Number"],
        "personalcode": "Aojfiwjisf" + ": "+str(data1["personal_code"])
    }

    # Сохраняем обновленные данные в JSON файл
    data1["personal_code"]+=1
    save_data(data1)
    await message.answer(f"Ваш персональный код: <code><b>{data1[str(user_id)]["personalcode"]}</b></code>",parse_mode="HTML")

    filename = "example.xlsx"

    def save_to_excel():
        # Загрузка данных из JSON-файла

        # Проверка существует ли файл
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Запись заголовков
            ws.append(["user_id", "personalcode", "phone_number"])

        # Создание множества для проверки уникальности
        existing_user_ids = set()
        existing_personalcodes = set()

        # Загрузка существующих данных из Excel-файла в множества
        for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            existing_user_ids.add(row[0])
            existing_personalcodes.add(row[1])  # personalcode находится во втором столбце (индекс 1)

        # Добавление новых данных, если они уникальны
        new_data_added = False
        for user_id, details in data1.items():
            if user_id != "personal_code":  # Игнорируем ключ "personal_code"
                personalcode = details["personalcode"]
                phone_number = details["phone_number"]

                if user_id not in existing_user_ids and personalcode not in existing_personalcodes:
                    ws.append([user_id, personalcode, phone_number])
                    existing_user_ids.add(user_id)
                    existing_personalcodes.add(personalcode)
                    new_data_added = True

        # Сохранение файла только если добавлены новые данные
        if new_data_added:
            wb.save(filename)
            print(f"Новые данные успешно сохранены в {filename}")
        else:
            print("Новых данных нет")

    save_to_excel()
    #json_string = json.dumps(data1, ensure_ascii=False, indent=4)
    #
    #await bot.send_message(
    #            chat_id=admin_id,
    #            text=json_string
    #        )





# Команда /профиль 
@dp.message(F.text.contains("Профиль"))
async def start(message: Message):
    data1 = load_data()
    await message.answer(f"""Профиль

<b>Имя</b>: {data1[str(message.from_user.id)]["name"]}
<b>Фамилия</b>: {data1[str(message.from_user.id)]["surname"]}
<b>Номер телефона</b>: {data1[str(message.from_user.id)]["phone_number"]}
<b>Персональный код</b>: {data1[str(message.from_user.id)]["personalcode"]}
""",parse_mode="HTML")
    await message.answer("Все верно?",reply_markup=buttons.vseverno)

#Все КОЛБЕКИ
@dp.callback_query(lambda callback_query: callback_query.data == "yes")
async def no(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Отлично!")

##################################################################################################################################################################################################
@dp.callback_query(lambda callback_query: callback_query.data == "net")
async def no(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Ок, что будем редактировать?",reply_markup=buttons.options)
#Callback Name 
@dp.callback_query(lambda callback_query: callback_query.data == "name")
async def no(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Введи Имя")
    await state.set_state(Change.newname)

@dp.message(Change.newname)
async def getnewname(message: Message,state: FSMContext):
    data = load_data()
    newname = message.text
    data[str(message.from_user.id)]['name'] = newname
    save_data(data)
    await message.answer(f"""Профиль

<b>Имя</b>: {data[str(message.from_user.id)]["name"]}
<b>Фамилия</b>: {data[str(message.from_user.id)]["surname"]}
<b>Номер телефона</b>: {data[str(message.from_user.id)]["phone_number"]}
""",parse_mode="HTML")
    await message.answer("Все верно?",reply_markup=buttons.vseverno)
    await state.clear()

#Callback Number 
@dp.callback_query(lambda callback_query: callback_query.data == "number")
async def no(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Введи номер")
    await state.set_state(Change.newnumber)

@dp.message(Change.newnumber)
async def getnewname(message: Message,state: FSMContext):
    data = load_data()
    newnumber = message.text
    data[str(message.from_user.id)]['phone_number'] = newnumber
    save_data(data)
    await message.answer(f"""Профиль

<b>Имя</b>: {data[str(message.from_user.id)]["name"]}
<b>Фамилия</b>: {data[str(message.from_user.id)]["surname"]}
<b>Номер телефона</b>: {data[str(message.from_user.id)]["phone_number"]}
""",parse_mode="HTML")
    await message.answer("Все верно?",reply_markup=buttons.vseverno)
    await state.clear()

@dp.callback_query(lambda callback_query: callback_query.data == "surname")
async def no(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Введи Фамилию")
    await state.set_state(Change.newsurname)

@dp.message(Change.newsurname)
async def getnewsurname(message: Message, state: FSMContext):
    data = load_data()
    newsurname = message.text
    data[str(message.from_user.id)]['surname'] = newsurname
    save_data(data)
    await message.answer(f"""Профиль

<b>Имя</b>: {data[str(message.from_user.id)]["name"]}
<b>Фамилия</b>: {data[str(message.from_user.id)]["surname"]}
<b>Номер телефона</b>: {data[str(message.from_user.id)]["phone_number"]}
""",parse_mode="HTML")
    await message.answer("Все верно?",reply_markup=buttons.vseverno)
    await state.clear()


#Инструкции 
@dp.message(F.text.contains("Инструкция"))
async def instruc(message: Message):
    await message.answer("Выберите маркетплейс:",reply_markup=buttons.marketplaces)
#пиндодо
@dp.callback_query(lambda callback_query: callback_query.data == "pd")
async def pd(callback: CallbackQuery, state: FSMContext):
    md = "BQACAgIAAxkBAAPlZl2y4lHRecyHSn8P1fQjqqWWiTYAAuZIAAKzxfBK2UzLnkW5U_41BA"
    await callback.message.reply_document(md)
    
@dp.callback_query(lambda callback_query: callback_query.data == "to")
async def to(callback: CallbackQuery, state: FSMContext):
    md = "BQACAgIAAxkBAAPnZl2y4rOdyv-fNslmg-SywdYF0WYAAuhIAAKzxfBKYPA-CxNbNY01BA"
    await callback.message.reply_document(md)

@dp.callback_query(lambda callback_query: callback_query.data == "18")
async def pd18(callback: CallbackQuery, state: FSMContext):
    md = "BQACAgIAAxkBAAPpZl2y4nuyZ9x2qhU1F-tcoICfRyIAAulIAAKzxfBKruMDkgYg5Oc1BA"
    await callback.message.reply_document(md)

@dp.callback_query(lambda callback_query: callback_query.data == "pn")
async def pn(callback: CallbackQuery, state: FSMContext):
    md = "BQACAgIAAxkBAAPrZl2y4vcSpsuOXxHUsZFJoRSli8YAAupIAAKzxfBKUA7mZqT4XLo1BA"
    await callback.message.reply_document(md)


# Поддержка 
@dp.message(F.text.contains("Поддержка"))
async def helpfromsky(message: Message):
    await message.answer("""⚙️Контакты поддержки⚔️
            <code>0550223344</code>""",parse_mode="HTML")
    
@dp.message(F.text.contains("Адреса"))
async def adresses(message: Message):
    data = load_data()
    await message.answer("""<code>MK- 
14757883517
 浙江省 金华市 义乌市
后宅街道群英路281号厂房二房屋一楼中俄进出口  MK- </code>""",parse_mode="HTML")


#Отмена калькулятора
@dp.message(F.text.contains("Отмена"))
async def notc(message: Message , state: FSMContext):
    await message.answer("Вы отменили калькулятор!",reply_markup=buttons.mainkb)   
    await state.clear()
    return 
    
#Менюшка после обычного запроса по тексту 
@dp.message(F.text.lower() == "меню")
async def mjogj(message: Message):
    await message.answer("<i><b>меню</b></i>",parse_mode="HTML",reply_markup=buttons.mainkb)




#Общий калькулятор 
@dp.message(F.text.contains("Калькулятор"))
async def calcs(message: Message,state: FSMContext):
    await message.answer("Введите длину (см): ",reply_markup=buttons.notkb)
    await state.set_state(Calc.length)

@dp.message(Calc.length)
async def getlength(message: Message, state: FSMContext):
    await state.update_data(length=message.text)
    await state.set_state(Calc.width)
    await message.answer("Введите ширину (см): ")

@dp.message(Calc.width)
async def getwidth(message: Message, state: FSMContext):
    await state.update_data(width=message.text)
    await state.set_state(Calc.height)
    await message.answer("Введите высоту (см): ")

@dp.message(Calc.height)
async def getheight(message: Message,state: FSMContext):
    await state.update_data(height=message.text)
    await state.set_state(Calc.kilos)
    await message.answer("Введите вес (кг): ")

@dp.message(Calc.kilos)
async def getkilos(message: Message, state: FSMContext):
    await state.update_data(kilos=message.text)
    data = await state.get_data()
    l1 = data["length"]
    w1 = data["width"]
    h1 = data["height"]
    k1 = data["kilos"]
    volume  = (float(l1) * float(w1) * float(h1) ) /   (float(k1) * 1000)
    await message.answer("Объемный вес: " + f"<code>{str(volume)}</code>"+ "кг",parse_mode="HTML",reply_markup=buttons.mainkb)
    await state.clear()


    #data1 = load_data()
    #keys = [key for key in data1]
    #true_keys = [i for i in keys if i.isdigit()]
@dp.message(Command("sendall"))
async def sendmessagestoall(message: Message,state: FSMContext):
    if str(message.from_user.id) == admin_id:
        await message.answer(f'Здравсвтуйте <b>{message.from_user.first_name}</b> . Хотите ли вы разослать сообщение с фоткой?',parse_mode="HTML",reply_markup=buttons.photoyesorno)
    else:
        await message.answer("Error!")

@dp.callback_query(lambda c: c.data in ["yesphoto", "nophoto"])
async def process_photo_choice(callback: CallbackQuery, state: FSMContext):
    if callback.data ==   "yesphoto":
        await callback.message.answer("Отправьте фотку")
        await state.set_state(sendall.getphoto)
    else:
        await callback.message.answer("Теперь отправьте основной текст!")
        await state.set_state(sendall.gettext)

@dp.message(sendall.getphoto , F.photo)
async def get_photo2(message: Message, state : FSMContext):
    await state.update_data(getphoto=message.photo[-1].file_id)
    await message.answer("Теперь отправьте основной текст!")
    await state.set_state(sendall.gettext)

@dp.message(sendall.gettext)
async def gettext123(message: Message, state: FSMContext):
    await state.update_data(gettext=message.text)
    await message.answer("Продолжить с кнопкой или нет ? ",reply_markup=buttons.buttonyesorno)

# Обработка выбора с кнопкой или без
@dp.callback_query(lambda c: c.data in ["yesbutton", "nobutton"])
async def process_button_choice(callback: CallbackQuery, state: FSMContext):
    if callback.data == "yesbutton":
        await callback.message.answer("Отправьте текст кнопки!")
        await state.set_state(sendall.getbutton)
    else:
        await send_preview(callback.message, state)

# Предпросмотр сообщения
async def send_preview(message, state):
    data = await state.get_data()
    if 'getphoto' in data and data['getphoto']:
        await bot.send_photo(
            chat_id=message.chat.id,
            photo=data['getphoto'],
            caption=data['gettext'],
            reply_markup=create_keyboard(data)
        )
    else:
        await message.answer(
            text=data['gettext'],
            reply_markup=create_keyboard(data)
        )
    await message.answer("Начинаем отправлять?", reply_markup=buttons.confirmationyesorno)

@dp.message(sendall.getbutton)
async def getbutton123(message: Message, state: FSMContext):
    await state.update_data(getbutton=message.text)
    await message.answer("Теперь отправьте ссылку, которую кнопка будет содержать:")
    await state.set_state(sendall.getlink)

@dp.message(sendall.getlink)
async def getlink342(message: Message, state: FSMContext):
    await state.update_data(getlink=message.text)
    await send_preview(message, state)





# Создание клавиатуры
def create_keyboard(data):
    if 'getbutton' in data and data['getbutton'] and 'getlink' in data and data['getlink']:
        return InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=data['getbutton'], url=data['getlink'])]
            ],
            resize_keyboard=True
        )
    return None


# Подтверждение рассылки
@dp.callback_query(lambda c: c.data == "yesconfirm")
async def confirm_send(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    added_keyboards = create_keyboard(data)

    data1 = load_data()
    keys = [key for key in data1]
    true_keys = [i for i in keys if i.isdigit()]
    j = 0
    for i in true_keys:
        if 'getphoto' in data and data['getphoto']:
            await bot.send_photo(
                chat_id=i,
                photo=data['getphoto'],
                caption=data['gettext'],
                reply_markup=added_keyboards
            )
            time.sleep(0.33)
        else:
            await bot.send_message(
                chat_id=i,
                text=data['gettext'],
                reply_markup=added_keyboards
            )
            time.sleep(0.33)
        j += 1
    await callback.message.answer(f"Количество отправленных рассылок: {j}")
    await state.clear()











    


    





    
    





async def main() -> None:
    
    # Initialize Bot instance with a default parse mode which will be passed to all API calls
    
    # And the run events dispatching
    await dp.start_polling(bot)

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, stream=sys.stdout)
    asyncio.run(main())