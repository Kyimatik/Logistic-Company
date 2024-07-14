import json
from openpyxl import Workbook, load_workbook
from main import load_data  # Импорт функции загрузки данных
import os
import schedule
import time

# Имя файла Excel
filename = "example.xlsx"

def save_to_excel():
    # Загрузка данных из JSON-файла
    data = load_data()

    # Проверка существует ли файл
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Запись заголовков
        ws.append(["user_id", "personalcode", "phone_number"])

    # Создание множества для проверки уникальности
    existing_user_ids = set()
    existing_personalcodes = set()

    # Загрузка существующих данных из Excel-файла в множества
    for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
        existing_user_ids.add(row[0])
        existing_personalcodes.add(row[1])  # personalcode находится во втором столбце (индекс 1)

    # Добавление новых данных, если они уникальны
    new_data_added = False
    for user_id, details in data.items():
        if user_id != "personal_code":  # Игнорируем ключ "personal_code"
            personalcode = details["personalcode"]
            phone_number = details["phone_number"]
            
            if user_id not in existing_user_ids and personalcode not in existing_personalcodes:
                ws.append([user_id, personalcode, phone_number])
                existing_user_ids.add(user_id)
                existing_personalcodes.add(personalcode)
                new_data_added = True

    # Сохранение файла только если добавлены новые данные
    if new_data_added:
        wb.save(filename)
        print(f"Новые данные успешно сохранены в {filename}")
    else:
        print("Новых данных нет")

# Планирование задачи на каждые 10 минут
schedule.every(1).minutes.do(save_to_excel)

# Начало выполнения планировщика
print("Начало выполнения планировщика. Для остановки нажмите Ctrl+C")
while True:
    schedule.run_pending()
    time.sleep(1)

